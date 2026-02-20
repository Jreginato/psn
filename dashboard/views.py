from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from django.template.loader import render_to_string
from django.utils import timezone
from produtos.models import AcessoProduto, Produto, Pedido


@login_required
def home(request):
    """Dashboard principal do cliente"""
    # Produtos com acesso
    acessos_qs = AcessoProduto.objects.filter(
        usuario=request.user,
        ativo=True
    ).select_related('produto', 'produto__categoria')
    acessos = acessos_qs.order_by('-liberado_em')
    
    # Últimos pedidos
    pedidos = Pedido.objects.filter(usuario=request.user).order_by('-criado_em')[:5]
    
    # Estatísticas
    total_produtos = acessos_qs.count()
    total_pedidos = pedidos.count()
    
    context = {
        'acessos': acessos,
        'pedidos': pedidos,
        'total_produtos': total_produtos,
        'total_pedidos': total_pedidos,
        'user_produtos': acessos,  # Para menu sidebar
    }
    return render(request, 'dashboard/home.html', context)


@login_required
def meus_produtos(request):
    """Lista todos os produtos do usuário"""
    acessos = AcessoProduto.objects.filter(
        usuario=request.user
    ).select_related('produto', 'produto__categoria').order_by('-liberado_em')
    
    context = {
        'acessos': acessos,
        'user_produtos': acessos[:10],  # Para menu sidebar
    }
    return render(request, 'dashboard/meus_produtos.html', context)


@login_required
def produto_detalhes(request, slug):
    """Visualizar conteúdo de um produto"""
    produto = get_object_or_404(Produto, slug=slug)
    
    # Verificar se tem acesso
    acesso = get_object_or_404(
        AcessoProduto,
        usuario=request.user,
        produto=produto
    )
    
    if not acesso.is_ativo:
        messages.error(request, 'Seu acesso a este produto expirou.')
        return redirect('dashboard:meus_produtos')
    
    # Atualizar último acesso e contador
    from django.utils import timezone
    acesso.ultimo_acesso = timezone.now()
    acesso.total_acessos += 1
    acesso.save()
    
    # Pegar conteúdos
    conteudos = produto.conteudos.filter(liberado=True).order_by('ordem')
    
    # Carregar produtos para sidebar
    user_produtos = AcessoProduto.objects.filter(
        usuario=request.user,
        ativo=True
    ).select_related('produto')[:10]
    
    context = {
        'produto': produto,
        'conteudos': conteudos,
        'acesso': acesso,
        'user_produtos': user_produtos,
    }
    return render(request, 'dashboard/produto_conteudo.html', context)


@login_required
def pedidos(request):
    """Histórico de pedidos"""
    pedidos_list = Pedido.objects.filter(
        usuario=request.user
    ).prefetch_related('itens').order_by('-criado_em')
    
    # Carregar produtos para sidebar
    user_produtos = AcessoProduto.objects.filter(
        usuario=request.user,
        ativo=True
    ).select_related('produto')[:10]
    
    context = {
        'pedidos': pedidos_list,
        'user_produtos': user_produtos,
    }
    return render(request, 'dashboard/pedidos.html', context)


@login_required
def perfil(request):
    """Perfil do usuário"""
    # Carregar produtos para sidebar
    user_produtos = AcessoProduto.objects.filter(
        usuario=request.user,
        ativo=True
    ).select_related('produto')[:10]
    
    context = {
        'user_produtos': user_produtos,
    }
    return render(request, 'dashboard/perfil.html', context)


# ===== CONSULTORIA ONLINE =====

@login_required
def consultoria_treino(request):
    """Área de treino da consultoria online"""
    if not request.user.tem_consultoria_online():
        messages.error(request, 'Você não tem acesso à consultoria online.')
        return redirect('dashboard:home')
    
    consultoria = request.user.consultoria
    treino_atual = consultoria.get_treino_atual()
    
    user_produtos = AcessoProduto.objects.filter(
        usuario=request.user,
        ativo=True
    ).select_related('produto')[:10]
    
    context = {
        'consultoria': consultoria,
        'treino_atual': treino_atual,
        'user_produtos': user_produtos,
    }
    return render(request, 'dashboard/consultoria/treino.html', context)


@login_required
def consultoria_treino_pdf(request):
    """Exporta o treino atual em PDF"""
    if not request.user.tem_consultoria_online():
        messages.error(request, 'Você não tem acesso à consultoria online.')
        return redirect('dashboard:home')

    consultoria = request.user.consultoria
    treino_atual = consultoria.get_treino_atual()

    if not treino_atual:
        messages.error(request, 'Nenhum treino atual para exportar.')
        return redirect('dashboard:consultoria_treino')

    try:
        from xhtml2pdf import pisa
    except ImportError:
        messages.error(request, 'Biblioteca de PDF nao instalada. Instale xhtml2pdf.')
        return redirect('dashboard:consultoria_treino')

    html_string = render_to_string(
        'dashboard/consultoria/treino_export_pdf.html',
        {
            'consultoria': consultoria,
            'treino_atual': treino_atual,
            'usuario': request.user,
            'gerado_em': timezone.now(),
        }
    )

    filename = f"treino_{request.user.username}.pdf"
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    pisa_status = pisa.CreatePDF(html_string, dest=response, encoding='utf-8')
    if pisa_status.err:
        messages.error(request, 'Nao foi possivel gerar o PDF.')
        return redirect('dashboard:consultoria_treino')
    return response


@login_required
def consultoria_treino_excel(request):
    """Exporta o treino atual em Excel"""
    if not request.user.tem_consultoria_online():
        messages.error(request, 'Você não tem acesso à consultoria online.')
        return redirect('dashboard:home')

    consultoria = request.user.consultoria
    treino_atual = consultoria.get_treino_atual()

    if not treino_atual:
        messages.error(request, 'Nenhum treino atual para exportar.')
        return redirect('dashboard:consultoria_treino')

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        messages.error(request, 'Biblioteca de Excel nao instalada. Instale openpyxl.')
        return redirect('dashboard:consultoria_treino')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Treino Atual'

    brand_font = Font(bold=True, size=14)
    header_font = Font(bold=True, size=12)
    muted_fill = PatternFill('solid', fgColor='EEF2F6')

    ws.merge_cells('A1:F1')
    ws['A1'] = 'PersonalTrainer'
    ws['A1'].font = brand_font
    ws['A1'].alignment = Alignment(horizontal='left')

    ws['A3'] = 'Aluno:'
    ws['B3'] = request.user.get_full_name() or request.user.username
    ws['A4'] = 'Treino:'
    ws['B4'] = treino_atual.titulo
    ws['A5'] = 'Gerado em:'
    ws['B5'] = timezone.now().strftime('%d/%m/%Y %H:%M')

    row = 7
    for dia in treino_atual.dias.all():
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.cell(row=row, column=1, value=dia.nome).font = header_font
        ws.cell(row=row, column=1).fill = muted_fill
        row += 1

        ws.cell(row=row, column=1, value='Exercicio').font = header_font
        ws.cell(row=row, column=2, value='Series').font = header_font
        ws.cell(row=row, column=3, value='Repeticoes').font = header_font
        ws.cell(row=row, column=4, value='Carga').font = header_font
        ws.cell(row=row, column=5, value='Descanso').font = header_font
        ws.cell(row=row, column=6, value='Obs').font = header_font
        ws.cell(row=row, column=7, value='Video').font = header_font
        row += 1

        for item in dia.exercicios.all():
            ws.cell(row=row, column=1, value=item.exercicio.nome)
            ws.cell(row=row, column=2, value=item.series)
            ws.cell(row=row, column=3, value=item.repeticoes)
            ws.cell(row=row, column=4, value=item.carga or '')
            ws.cell(row=row, column=5, value=item.descanso or '')
            ws.cell(row=row, column=6, value=item.observacao_especifica or '')
            ws.cell(row=row, column=7, value=item.exercicio.link_video or '')
            row += 1

        row += 1

    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[col].width = 22

    filename = f"treino_{request.user.username}.xlsx"
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
def consultoria_plano_alimentar(request):
    """Área de plano alimentar da consultoria online"""
    if not request.user.tem_consultoria_online():
        messages.error(request, 'Você não tem acesso à consultoria online.')
        return redirect('dashboard:home')
    
    consultoria = request.user.consultoria
    user_produtos = AcessoProduto.objects.filter(
        usuario=request.user,
        ativo=True
    ).select_related('produto')[:10]
    
    context = {
        'consultoria': consultoria,
        'user_produtos': user_produtos,
    }
    return render(request, 'dashboard/consultoria/plano_alimentar.html', context)


@login_required
def consultoria_medicacao(request):
    """Área de medicação da consultoria online"""
    if not request.user.tem_consultoria_online():
        messages.error(request, 'Você não tem acesso à consultoria online.')
        return redirect('dashboard:home')
    
    consultoria = request.user.consultoria
    user_produtos = AcessoProduto.objects.filter(
        usuario=request.user,
        ativo=True
    ).select_related('produto')[:10]
    
    context = {
        'consultoria': consultoria,
        'user_produtos': user_produtos,
    }
    return render(request, 'dashboard/consultoria/medicacao.html', context)


